"""
Procesador de Lotes Mejorado
Procesa carpetas completas con generación de reportes PDF/Excel
"""

from core.clasif icador import ClasificadorLimones
from core.config_manager import ConfigManager
from core.grupos_manager import GruposManager
from database.dao import LimonDAO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os
import cv2
from typing import List, Dict
from PIL import Image
import io


class BatchProcessor:
    """
    Procesa lotes de imágenes y genera reportes.
    """
    
    def __init__(self, config_manager=None, grupos_manager=None):
        self.config = config_manager or ConfigManager()
        self.grupos_manager = grupos_manager or GruposManager()
        self.clasificador = ClasificadorLimones(config_manager=self.config, verbose=False)
        self.limon_dao = LimonDAO()
        
        self.resultados = []
    
    def procesar_carpeta(self, carpeta_path: str, guardar_bd: bool = True,
                        callback_progreso=None) -> List[Dict]:
        """
        Procesar todas las imágenes en una carpeta.
        
        Args:
            carpeta_path: Ruta a la carpeta
            guardar_bd: Si debe guardar en base de datos
            callback_progreso: Función para reportar progreso (idx, total)
        
        Returns:
            Lista de resultados
        """
        self.resultados = []
        
        # Obtener imágenes
        extensiones = ['.jpg', '.

jpeg', '.png', '.bmp']
        imagenes = [f for f in os.listdir(carpeta_path) 
                   if os.path.splitext(f)[1].lower() in extensiones]
        
        total = len(imagenes)
        
        for idx, filename in enumerate(imagenes):
            try:
                ruta_completa = os.path.join(carpeta_path, filename)
                
                # Clasificar
                self.clasificador.cargar_imagen(ruta_completa)
                resultado = self.clasificador.procesar(mostrar_visualizacion=False)
                
                if resultado:
                    vector = resultado['vector_caracteristicas']
                    acidez = resultado['acidez_estimada']
                    grupos = self.grupos_manager.clasificar_multi_grupo(vector, acidez)
                    grupo_final = self.grupos_manager.obtener_grupo_prioritario(grupos)
                    salida = self.grupos_manager.obtener_salida_fisica(grupo_final)
                    
                    resultado_completo = {
                        'archivo': filename,
                        'ruta': ruta_completa,
                        'tonalidad': vector[0],
                        'rugosidad': vector[1],
                        'defectos': vector[2],
                        'acidez': acidez,
                        'grupo': grupo_final,
                        'salida': salida,
                        'clasificacion_original': resultado['clasificacion'],
                        'vida_util': resultado['vida_util_dias']
                    }
                    
                    self.resultados.append(resultado_completo)
                    
                    # Guardar en BD
                    if guardar_bd:
                        self.limon_dao.guardar(
                            tonalidad=vector[0],
                            rugosidad=vector[1],
                            defectos=vector[2],
                            acidez=acidez,
                            grupo_asignado=grupo_final,
                            salida_fisica=salida,
                            clasificacion_original=resultado['clasificacion'],
                            vida_util_dias=resultado['vida_util_dias'],
                            imagen_path=ruta_completa,
                            modo='lote'
                        )
                
                # Callback de progreso
                if callback_progreso:
                    callback_progreso(idx + 1, total)
                    
            except Exception as e:
                print(f"Error procesando {filename}: {e}")
        
        return self.resultados
    
    def generar_pdf(self, output_path: str):
        """Generar reporte PDF"""
        doc = SimpleDocTemplate(output_path, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo = Paragraph(f"<b>Reporte de Lote de Clasificación</b><br/>{datetime.now().strftime('%Y-%m-%d %H:%M')}", 
                          styles['Title'])
        elements.append(titulo)
        elements.append(Spacer(1, 0.3*inch))
        
        # Resumen
        total = len(self.resultados)
        conteo_grupos = {}
        for r in self.resultados:
            conteo_grupos[r['grupo']] = conteo_grupos.get(r['grupo'], 0) + 1
        
        resumen_text = f"<b>Total procesados:</b> {total}<br/>"
        for grupo, count in sorted(conteo_grupos.items(), key=lambda x: x[1], reverse=True):
            porcentaje = (count/total*100) if total > 0 else 0
            resumen_text += f"• {grupo}: {count} ({porcentaje:.1f}%)<br/>"
        
        resumen = Paragraph(resumen_text, styles['Normal'])
        elements.append(resumen)
        elements.append(Spacer(1, 0.3*inch))
        
        # Tabla de resultados
        data = [['Archivo', 'Grupo', 'Acidez %', 'Rugosidad', 'Defectos %', 'Salida']]
        
        for r in self.resultados[:50]:  # Limitar a 50 para no saturar PDF
            data.append([
                r['archivo'][:30],
                r['grupo'],
                f"{r['acidez']:.1f}",
                f"{r['rugosidad']:.1f}",
                f"{r['defectos']:.1f}",
                str(r['salida'])
            ])
        
        tabla = Table(data)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(tabla)
        
        # Generar PDF
        doc.build(elements)
        print(f"✓ PDF generado: {output_path}")
    
    def generar_excel(self, output_path: str):
        """Generar reporte Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        
        # Headers
        headers = ['Archivo', 'Grupo', 'Tonalidad', 'Rugosidad', 'Defectos', 
                  'Acidez', 'Vida Útil', 'Salida', 'Clasificación Original']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, resultado in enumerate(self.resultados, 2):
            ws.cell(row=row_idx, column=1, value=resultado['archivo'])
            ws.cell(row=row_idx, column=2, value=resultado['grupo'])
            ws.cell(row=row_idx, column=3, value=round(resultado['tonalidad'], 2))
            ws.cell(row=row_idx, column=4, value=round(resultado['rugosidad'], 2))
            ws.cell(row=row_idx, column=5, value=round(resultado['defectos'], 2))
            ws.cell(row=row_idx, column=6, value=round(resultado['acidez'], 2))
            ws.cell(row=row_idx, column=7, value=resultado['vida_util'])
            ws.cell(row=row_idx, column=8, value=resultado['salida'])
            ws.cell(row=row_idx, column=9, value=resultado['clasificacion_original'])
        
        # Ajustar anchos
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 40)
        
        # Hoja de resumen
        ws_resumen = wb.create_sheet("Resumen")
        ws_resumen.cell(1, 1, "Grupo").font = Font(bold=True)
        ws_resumen.cell(1, 2, "Total").font = Font(bold=True)
        ws_resumen.cell(1, 3, "Porcentaje").font = Font(bold=True)
        
        conteo_grupos = {}
        for r in self.resultados:
            conteo_grupos[r['grupo']] = conteo_grupos.get(r['grupo'], 0) + 1
        
        total = len(self.resultados)
        row = 2
        for grupo, count in sorted(conteo_grupos.items(), key=lambda x: x[1], reverse=True):
            ws_resumen.cell(row, 1, grupo)
            ws_resumen.cell(row, 2, count)
            ws_resumen.cell(row, 3, f"{(count/total*100):.1f}%")
            row += 1
        
        # Guardar
        wb.save(output_path)
        print(f"✓ Excel generado: {output_path}")


if __name__ == "__main__":
    print("=== Test de Batch Processor ===\n")
    
    processor = BatchProcessor()
    
    # Crear carpeta de prueba si no existe
    test_dir = "test_lote"
    if not os.path.exists(test_dir):
        print("⚠️ Carpeta de prueba no existe. Copia algunas imágenes a 'test_lote/' para probar.")
    else:
        print("1. Procesando carpeta...")
        resultados = processor.procesar_carpeta(test_dir, guardar_bd=False)
        print(f"   ✓ {len(resultados)} imágenes procesadas")
        
        if resultados:
            print("\n2. Generando PDF...")
            processor.generar_pdf("reporte_lote.pdf")
            
            print("\n3. Generando Excel...")
            processor.generar_excel("reporte_lote.xlsx")
            
            print("\n✅ Reportes generados")
        else:
            print("\n⚠️ No se procesaron imágenes")
